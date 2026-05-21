import sys
import os
import threading
import re
import shutil
from datetime import datetime
from flask import Flask, request, jsonify
import pandas as pd
import openpyxl
import webview

app = Flask(__name__)

# 재고 DB 파일 이름 및 폴더 설정
DB_FILE = '부품_재고현황.xlsx'
SAVE_FOLDER = '납품 명세서'
PRICE_FILE = '단가.xlsx'
SALES_FILE = '매출.xlsx'

def get_price_file():
    # 작업 디렉토리에 단가.xlsx가 없으면 dist 또는 Backup 폴더에서 가져옵니다.
    if not os.path.exists(PRICE_FILE):
        dist_path = os.path.join('dist', PRICE_FILE)
        backup_path = os.path.join('STMS Backup', PRICE_FILE)
        if os.path.exists(dist_path):
            shutil.copy(dist_path, PRICE_FILE)
        elif os.path.exists(backup_path):
            shutil.copy(backup_path, PRICE_FILE)
        else:
            raise FileNotFoundError("단가.xlsx 파일을 찾을 수 없습니다. STMS_code 폴더에 단가.xlsx 파일을 추가해 주세요.")

def calculate_sales():
    get_price_file()
    
    if not os.path.exists(SAVE_FOLDER):
        os.makedirs(SAVE_FOLDER)
        
    csv_files = [f for f in os.listdir(SAVE_FOLDER) if f.endswith('.csv')]
    if not csv_files:
        return None
        
    price_df = pd.read_excel(PRICE_FILE)
    price_df['품번'] = price_df['품번'].astype(str).str.strip()
    price_dict = dict(zip(price_df['품번'], price_df['단가']))
    price_name_dict = dict(zip(price_df['품번'], price_df['품명']))
    
    all_records = []
    for csv_file in csv_files:
        csv_path = os.path.join(SAVE_FOLDER, csv_file)
        
        # 날짜 정규식 파싱 시도 (실패 시 파일 수정일로 폴백)
        match = re.search(r'_(\d{8})\.csv$', csv_file)
        if match:
            date_str = match.group(1)
            formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
        else:
            try:
                mtime = os.path.getmtime(csv_path)
                formatted_date = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d')
            except Exception:
                continue
        
        # 다중 인코딩 시도로 안전하게 CSV 파일 로드
        df = None
        for enc in ['utf-8-sig', 'cp949', 'utf-8', 'euc-kr']:
            try:
                df = pd.read_csv(csv_path, encoding=enc)
                break
            except Exception:
                continue
                
        if df is None:
            print(f"[Error] 인코딩 오류로 CSV 파일을 읽을 수 없습니다: {csv_file}")
            continue
            
        df.columns = [str(c).strip() for c in df.columns]
        if '품번' not in df.columns or '납품수량' not in df.columns:
            continue
            
        for _, row in df.iterrows():
            p_no = str(row['품번']).strip()
            if not p_no or p_no == 'nan':
                continue
            try:
                qty = float(str(row['납품수량']).replace(',', ''))
            except ValueError:
                qty = 0.0
                
            unit_price = price_dict.get(p_no, 0)
            p_name = price_name_dict.get(p_no, str(row.get('품명', '')).strip())
            amount = unit_price * qty
            
            all_records.append({
                '날짜': formatted_date,
                '품번': p_no,
                '품명': p_name,
                '납품수량': qty,
                '단가': unit_price,
                '매출': amount
            })
            
    if not all_records:
        return None
        
    sales_df = pd.DataFrame(all_records)
    sales_df['날짜_dt'] = pd.to_datetime(sales_df['날짜'])
    
    # 1. 일일 품명별 매출
    daily_item = sales_df.groupby(['날짜', '품번', '품명'], as_index=False)[['납품수량', '매출']].sum()
    
    # 2. 일일 총 매출
    daily_total = sales_df.groupby('날짜', as_index=False)['매출'].sum()
    daily_total.rename(columns={'매출': '일일 총 매출'}, inplace=True)
    
    summary_df = sales_df.copy()
    
    # 3. 주간 총 매출
    summary_df['주차_시작'] = summary_df['날짜_dt'] - pd.to_timedelta(summary_df['날짜_dt'].dt.weekday, unit='D')
    summary_df['주차_끝'] = summary_df['주차_시작'] + pd.to_timedelta(6, unit='D')
    summary_df['주간'] = summary_df.apply(lambda r: f"{r['주차_시작'].strftime('%Y-%m-%d')} ~ {r['주차_끝'].strftime('%Y-%m-%d')}", axis=1)
    weekly_total = summary_df.groupby('주간', as_index=False)['매출'].sum()
    weekly_total.rename(columns={'매출': '주간 총 매출'}, inplace=True)
    
    # 4. 월간 총 매출
    summary_df['월간'] = summary_df['날짜_dt'].dt.strftime('%Y-%m')
    monthly_total = summary_df.groupby('월간', as_index=False)['매출'].sum()
    monthly_total.rename(columns={'매출': '월간 총 매출'}, inplace=True)
    
    # 5. 분기 총 매출
    summary_df['분기'] = summary_df['날짜_dt'].apply(lambda x: f"{x.year}-Q{(x.month-1)//3 + 1}")
    quarterly_total = summary_df.groupby('분기', as_index=False)['매출'].sum()
    quarterly_total.rename(columns={'매출': '분기 총 매출'}, inplace=True)
    
    # 6. 반기 총 매출
    summary_df['반기'] = summary_df['날짜_dt'].apply(lambda x: f"{x.year}-{(x.month-1)//6 + 1}H")
    half_yearly_total = summary_df.groupby('반기', as_index=False)['매출'].sum()
    half_yearly_total.rename(columns={'매출': '반기 총 매출'}, inplace=True)
    
    # 7. 연간 총 매출
    summary_df['연간'] = summary_df['날짜_dt'].dt.strftime('%Y')
    yearly_total = summary_df.groupby('연간', as_index=False)['매출'].sum()
    yearly_total.rename(columns={'매출': '연간 총 매출'}, inplace=True)
    
    # 엑셀 파일 저장
    with pd.ExcelWriter(SALES_FILE, engine='openpyxl') as writer:
        daily_item.to_excel(writer, sheet_name='일일 품명별 매출', index=False)
        daily_total.to_excel(writer, sheet_name='일일 총 매출', index=False)
        weekly_total.to_excel(writer, sheet_name='주간 총 매출', index=False)
        monthly_total.to_excel(writer, sheet_name='월간 총 매출', index=False)
        quarterly_total.to_excel(writer, sheet_name='분기 총 매출', index=False)
        half_yearly_total.to_excel(writer, sheet_name='반기 총 매출', index=False)
        yearly_total.to_excel(writer, sheet_name='연간 총 매출', index=False)
        
    return {
        'daily_item': daily_item.to_dict(orient='records'),
        'daily_total': daily_total.to_dict(orient='records'),
        'weekly_total': weekly_total.to_dict(orient='records'),
        'monthly_total': monthly_total.to_dict(orient='records'),
        'quarterly_total': quarterly_total.to_dict(orient='records'),
        'half_yearly_total': half_yearly_total.to_dict(orient='records'),
        'yearly_total': yearly_total.to_dict(orient='records'),
        'inventory': parse_inventory()
    }

def parse_inventory(file_path=None):
    if file_path is None:
        file_path = 'dist/부품_재고현황.xlsx'
        if not os.path.exists(file_path):
            file_path = '부품_재고현황.xlsx'
        if not os.path.exists(file_path):
            backup_path = 'STMS Backup/부품_재고현황.xlsx'
            if os.path.exists(backup_path):
                try:
                    shutil.copy(backup_path, '부품_재고현황.xlsx')
                    file_path = '부품_재고현황.xlsx'
                except Exception:
                    pass
                
    if not os.path.exists(file_path):
        return []
        
    try:
        df = pd.read_excel(file_path, header=None)
        inventory_list = []
        current_freq = ""
        
        for idx, row in df.iterrows():
            val0 = str(row[0]).strip() if pd.notna(row[0]) else ""
            if "납품 빈도" in val0:
                current_freq = val0.replace("납품 빈도 :", "").strip()
                continue
                
            val1 = str(row[1]).strip() if pd.notna(row[1]) else ""
            if val1 == "신품번" or val1 == "" or val0 == "순번":
                continue
                
            p_no = val1
            try:
                qty_val = row[4]
                if pd.isna(qty_val) or str(qty_val).strip() == '-':
                    stock_qty = 0
                else:
                    stock_qty = int(qty_val)
            except Exception:
                stock_qty = 0
                
            inventory_list.append({
                "신품번": p_no,
                "납품빈도": current_freq,
                "재고수량": stock_qty
            })
        return inventory_list
    except Exception as e:
        print(f"인벤토리 파싱 에러: {e}")
        return []

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- 페이지 라우팅 ---

@app.route('/')
@app.route('/dashboard')
def dashboard_page():
    """대시보드 페이지 (초기 화면)"""
    path = resource_path('dashboard.html')
    try:
        return open(path, encoding='utf-8').read()
    except FileNotFoundError:
        return "dashboard.html 파일을 찾을 수 없습니다."

@app.route('/inventory_system')
def inventory_system():
    """기존 재고 관리 페이지"""
    path = resource_path('index.html')
    try:
        return open(path, encoding='utf-8').read()
    except FileNotFoundError:
        return "index.html 파일을 찾을 수 없습니다."


# --- 기존 데이터 처리 로직 ---

@app.route('/inventory', methods=['GET'])
def get_inventory():
    if not os.path.exists(DB_FILE):
        return jsonify([])
    try:
        # 3행 헤더 설정 [cite: 246]
        df = pd.read_excel(DB_FILE, header=2)
        df.columns = [str(c).strip() for c in df.columns]
        
        target_order = ['순번', '신품번', '구품번', '품명', '재고수량', '공정 진행중', '입고수량']
        cols = [c for c in target_order if c in df.columns]
        df = df[cols].fillna('')
        return jsonify(df.to_dict(orient='records'))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400
    
    file = request.files['file']
    filename = file.filename
    
    if not os.path.exists(SAVE_FOLDER):
        os.makedirs(SAVE_FOLDER)
    
    try:
        # CSV 변환 저장 [cite: 751, 756]
        delivery_df = pd.read_excel(file)
        today_str = datetime.now().strftime('%Y%m%d')
        base_name = os.path.splitext(filename)[0]
        csv_filename = f"{base_name}_{today_str}.csv"
        csv_path = os.path.join(SAVE_FOLDER, csv_filename)
        delivery_df.to_csv(csv_path, index=False, encoding='utf-8-sig')

        # 변경 이전 재고현황 백업 저장 (부품_재고현황.json 단일 파일에 누적)
        json_file_path = '부품_재고현황.json'
        before_inventory = parse_inventory(DB_FILE)
        
        import json
        try:
            history_data = {}
            if os.path.exists(json_file_path):
                try:
                    with open(json_file_path, 'r', encoding='utf-8') as f:
                        history_data = json.load(f)
                except Exception as ex:
                    print(f"[Backup] 기존 JSON 읽기 실패(신규 생성): {ex}")
                    history_data = {}
            
            today_int = int(today_str)
            for item in before_inventory:
                p_no = item.get('신품번')
                if not p_no:
                    continue
                stock_qty = item.get('재고수량', 0)
                
                if p_no not in history_data:
                    history_data[p_no] = []
                
                # 동일 날짜가 있는지 체크하여 덮어쓰거나 새로 추가
                existing_entry = next((x for x in history_data[p_no] if x.get('날짜') == today_int), None)
                if existing_entry:
                    existing_entry['재고수량'] = stock_qty
                else:
                    history_data[p_no].append({
                        "날짜": today_int,
                        "재고수량": stock_qty
                    })
            
            with open(json_file_path, 'w', encoding='utf-8') as f:
                json.dump(history_data, f, ensure_ascii=False, indent=4)
            print(f"[Backup] 부품_재고현황.json 누적 백업 완료")
        except Exception as e:
            print(f"[Backup Error] 부품_재고현황.json 백업 실패: {e}")

        # 재고 차감 로직 (서식 유지) [cite: 265, 512]
        wb = openpyxl.load_workbook(DB_FILE)
        ws = wb.active
        header_row_num = 3
        
        col_map = {str(ws.cell(row=header_row_num, column=i).value).strip(): i 
                   for i in range(1, ws.max_column + 1)}
        
        delivery_df.columns = [str(c).strip() for c in delivery_df.columns]
        summary = delivery_df.groupby('품번')['납품수량'].sum().reset_index()
        
        chart_data_today = []
        for _, row in summary.iterrows():
            p_no = str(row['품번']).strip()
            qty = row['납품수량']
            chart_data_today.append({"label": p_no, "value": qty})
            
            for r_idx in range(header_row_num + 1, ws.max_row + 1):
                cell_p_no = str(ws.cell(row=r_idx, column=col_map['신품번']).value).strip()
                if cell_p_no == p_no:
                    raw_val = ws.cell(row=r_idx, column=col_map['재고수량']).value
                    current_qty = float(str(raw_val).replace(',', '')) if raw_val else 0
                    ws.cell(row=r_idx, column=col_map['재고수량']).value = current_qty - qty
                    break
        
        wb.save(DB_FILE)
        
        inventory_df = pd.read_excel(DB_FILE, header=2)
        chart_data_total = inventory_df[['신품번', '재고수량']].head(10).to_dict(orient='records')

        return jsonify({
            "status": "success",
            "message": f"{csv_filename} 저장 및 재고 차감 완료",
            "lineData": chart_data_today,
            "barData": chart_data_total
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/open_folder', methods=['POST'])
def open_folder():
    if not os.path.exists(SAVE_FOLDER):
        os.makedirs(SAVE_FOLDER)
    try:
        os.startfile(os.path.abspath(SAVE_FOLDER))
        return jsonify({"status": "success", "message": "폴더가 열렸습니다."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/sales_data', methods=['GET'])
def get_sales_data():
    try:
        data = calculate_sales()
        if data is None:
            return jsonify({"error": "매출 계산에 필요한 CSV 데이터가 없습니다."}), 404
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/inventory_history', methods=['GET'])
def get_inventory_history():
    p_no = request.args.get('p_no', '').strip()
    if not p_no:
        return jsonify({"error": "품번(p_no) 매개변수가 필요합니다."}), 400
        
    json_file_path = '부품_재고현황.json'
    history = []
    
    import json
    if os.path.exists(json_file_path):
        try:
            with open(json_file_path, 'r', encoding='utf-8') as f:
                history_data = json.load(f)
            
            p_history = history_data.get(p_no, [])
            for item in p_history:
                date_int = item.get('날짜')
                qty = item.get('재고수량', 0)
                if date_int:
                    date_str = str(date_int)
                    if len(date_str) == 8:
                        formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
                        history.append({"날짜": formatted_date, "재고수량": qty})
        except Exception as e:
            print(f"부품_재고현황.json 파싱 에러: {e}")
            
    # 정렬 및 반환
    sorted_history = sorted(history, key=lambda x: x["날짜"])
    return jsonify(sorted_history)

@app.route('/open_sales_file', methods=['POST'])
def open_sales_file():
    try:
        if not os.path.exists(SALES_FILE):
            calculate_sales()
        if os.path.exists(SALES_FILE):
            os.startfile(os.path.abspath(SALES_FILE))
            return jsonify({"status": "success", "message": "매출.xlsx 파일이 열렸습니다."})
        else:
            return jsonify({"error": "매출.xlsx 파일이 존재하지 않고 계산할 데이터도 없습니다."}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/update_inventory', methods=['POST'])
def update_inventory():
    """수정 모드 저장 로직 [cite: 167, 766]"""
    try:
        updated_data = request.json
        wb = openpyxl.load_workbook(DB_FILE)
        ws = wb.active
        header_row_num = 3
        
        # ws.max_column 괄호 제거로 에러 방지 [cite: 623, 666]
        col_map = {str(ws.cell(row=header_row_num, column=i).value).strip(): i 
                   for i in range(1, ws.max_column + 1)}

        for item in updated_data:
            p_no = str(item.get('신품번')).strip()
            if not p_no or p_no == 'None': continue

            for r_idx in range(header_row_num + 1, ws.max_row + 1):
                cell_p_no = str(ws.cell(row=r_idx, column=col_map['신품번']).value).strip()
                if cell_p_no == p_no:
                    for col_name, value in item.items():
                        if col_name in col_map and col_name not in ['순번', '신품번']:
                            if col_name in ['재고수량', '공정 진행중', '입고수량']:
                                try:
                                    ws.cell(row=r_idx, column=col_map[col_name]).value = float(str(value).replace(',', ''))
                                except:
                                    ws.cell(row=r_idx, column=col_map[col_name]).value = 0
                            else:
                                ws.cell(row=r_idx, column=col_map[col_name]).value = value
                    break
        
        wb.save(DB_FILE)
        return jsonify({"status": "success", "message": "저장되었습니다."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def run_flask():
    app.run(host='127.0.0.1', port=5000)

if __name__ == '__main__':
    if os.environ.get('STMS_NO_WEBVIEW') == '1':
        print("Starting Flask server only (NO_WEBVIEW mode)...")
        app.run(host='127.0.0.1', port=5000)
    else:
        t = threading.Thread(target=run_flask)
        t.daemon = True
        t.start()
        # pywebview를 이용한 전용 창 실행 [cite: 365, 518]
        webview.create_window('STMS 재고 관리 시스템', 'http://127.0.0.1:5000', width=1200, height=900)
        webview.start()