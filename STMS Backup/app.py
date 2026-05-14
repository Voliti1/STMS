import sys
import os
import threading
from datetime import datetime
from flask import Flask, request, jsonify
import pandas as pd
import openpyxl
import webview

app = Flask(__name__)

# 재고 DB 파일 이름 및 폴더 설정
DB_FILE = '부품_재고현황.xlsx'
SAVE_FOLDER = '납품 명세서'

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

@app.route('/')
def index():
    path = resource_path('index.html')
    try:
        return open(path, encoding='utf-8').read()
    except FileNotFoundError:
        return "index.html 파일을 찾을 수 없습니다."

@app.route('/inventory', methods=['GET'])
def get_inventory():
    if not os.path.exists(DB_FILE):
        return jsonify([])
    try:
        # 3행 헤더 설정 [cite: 246]
        df = pd.read_excel(DB_FILE, header=2)
        df.columns = [str(c).strip() for c in df.columns]
        
        # 정해진 컬럼 순서 [cite: 530, 533]
        target_order = ['순번', '신품번', '구품번', '품명', '재고수량', '공정 진행중', '입고수량']
        cols = [c for c in target_order if c in df.columns]
        df = df[cols].fillna('')
        return jsonify(df.to_dict(orient='records'))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400
    
    file = request.files['file']
    filename = file.filename
    
    if not os.path.exists(SAVE_FOLDER):
        os.makedirs(SAVE_FOLDER)
    
    try:
        delivery_df = pd.read_excel(file)
        today_str = datetime.now().strftime('%Y%m%d')
        base_name = os.path.splitext(filename)[0]
        csv_filename = f"{base_name}_{today_str}.csv"
        csv_path = os.path.join(SAVE_FOLDER, csv_filename)
        delivery_df.to_csv(csv_path, index=False, encoding='utf-8-sig')

        if not os.path.exists(DB_FILE):
            return jsonify({"error": "재고현황 파일이 없습니다."}), 500
            
        wb = openpyxl.load_workbook(DB_FILE)
        ws = wb.active
        header_row_num = 3
        
        # ws.max_column은 속성이므로 괄호()를 붙이지 않습니다 
        col_map = {str(ws.cell(row=header_row_num, column=i).value).strip(): i 
                   for i in range(1, ws.max_column + 1)}
        
        delivery_df.columns = [str(c).strip() for c in delivery_df.columns]
        summary = delivery_df.groupby('품번')['납품수량'].sum().reset_index()
        
        chart_data_today = []
        
        for _, row in summary.iterrows():
            p_no = str(row['품번']).strip()
            qty = row['납품수량']
            chart_data_today.append({"label": p_no, "value": qty})
            
            for r_idx in range(header_row_num + 1, ws.max_row + 1):
                cell_p_no = str(ws.cell(row=r_idx, column=col_map['신품번']).value).strip()
                if cell_p_no == p_no:
                    raw_val = ws.cell(row=r_idx, column=col_map['재고수량']).value
                    current_qty = float(str(raw_val).replace(',', '')) if raw_val else 0
                    ws.cell(row=r_idx, column=col_map['재고수량']).value = current_qty - qty
                    break
        
        wb.save(DB_FILE)
        
        inventory_df = pd.read_excel(DB_FILE, header=2)
        chart_data_total = inventory_df[['신품번', '재고수량']].head(10).to_dict(orient='records')

        return jsonify({
            "status": "success",
            "message": f"{csv_filename} 저장 및 재고 차감 완료",
            "lineData": chart_data_today,
            "barData": chart_data_total
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# 누락되었던 수정 모드 저장 로직 추가 
@app.route('/update_inventory', methods=['POST'])
def update_inventory():
    try:
        updated_data = request.json
        if not updated_data:
            return jsonify({"error": "데이터가 없습니다."}), 400

        wb = openpyxl.load_workbook(DB_FILE)
        ws = wb.active
        header_row_num = 3
        
        # 컬럼 위치 맵핑 생성
        col_map = {str(ws.cell(row=header_row_num, column=i).value).strip(): i 
                   for i in range(1, ws.max_column + 1)}

        for item in updated_data:
            p_no = str(item.get('신품번')).strip()
            if not p_no or p_no == 'None': continue

            # 엑셀에서 해당 신품번 찾기
            for r_idx in range(header_row_num + 1, ws.max_row + 1):
                cell_p_no = str(ws.cell(row=r_idx, column=col_map['신품번']).value).strip()
                if cell_p_no == p_no:
                    # 변경 가능한 컬럼들 업데이트 [cite: 74, 94]
                    for col_name, value in item.items():
                        if col_name in col_map and col_name not in ['순번', '신품번']:
                            # 숫자형 컬럼은 숫자로 변환하여 저장 
                            if col_name in ['재고수량', '공정 진행중', '입고수량']:
                                try:
                                    ws.cell(row=r_idx, column=col_map[col_name]).value = float(str(value).replace(',', ''))
                                except:
                                    ws.cell(row=r_idx, column=col_map[col_name]).value = 0
                            else:
                                ws.cell(row=r_idx, column=col_map[col_name]).value = value
                    break
        
        wb.save(DB_FILE)
        return jsonify({"status": "success", "message": "재고 현황이 성공적으로 업데이트되었습니다."})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def run_flask():
    app.run(host='127.0.0.1', port=5000)

if __name__ == '__main__':
    t = threading.Thread(target=run_flask)
    t.daemon = True
    t.start()
    webview.create_window('STMS 재고 관리 시스템', 'http://127.0.0.1:5000', width=1200, height=900)
    webview.start()